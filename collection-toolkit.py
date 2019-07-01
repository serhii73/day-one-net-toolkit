#!/usr/bin/env python
# Author: Daniel Teycheney

# Import Modules
from nornir import InitNornir
from nornir.plugins.tasks.networking import napalm_get
from nornir.plugins.tasks.networking import napalm_cli
from nornir.plugins.tasks.files import write_file
from nornir.plugins.functions.text import print_result
import json
import requests
import pathlib
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import datetime as dt

# Disable urllib3 warnings
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
import openpyxl


# Functions


def get_facts(task):
    task.run(name="Get facts", task=napalm_get, getters=["facts"])
    return "Complete"


def get_interfaces(task):
    task.run(name="Get interfaces", task=napalm_get, getters=["interfaces"])
    return "Complete"


def get_interfaces_ip(task):
    task.run(name="Get interfaces IP", task=napalm_get, getters=["interfaces_ip"])
    return "Complete"


def get_lldp_neighbors(task):
    task.run(name="Get LLDP neighbors", task=napalm_get, getters=["lldp_neighbors"])
    return "Complete"


def main_collector(wb):
    # Create facts worksheet
    facts_ws = wb.create_sheet("Facts")
    # Statically assign headers
    facts_headers = [
        "Hostname",
        "Vendor",
        "Model",
        "OS Version",
        "Serial Number",
        "Uptime",
    ]
    # Write headers on the top line of the file
    facts_ws.append(facts_headers)
    # Create interfaces worksheet
    interfaces_ws = wb.create_sheet("Interfaces")
    # Statically assign headers
    interfaces_headers = [
        "Name",
        "Interface Name",
        "Interface Description",
        "Interface Up",
        "Interface Enabled",
    ]
    # Write headers on the top line of the file
    interfaces_ws.append(interfaces_headers)
    # Create interfaces IP worksheet
    interfaces_ip_ws = wb.create_sheet("Interfaces_IP")
    # Statically assign headers
    interfaces_ip_headers = [
        "Name",
        "Interface Name",
        "IPv4 Address",
        "IPv4 Prefix Length",
    ]
    # Write headers on the top line of the file
    interfaces_ip_ws.append(interfaces_ip_headers)
    # Create LLDP neighbors worksheet
    lldp_nei_ws = wb.create_sheet("LLDP_Neighbor")
    # Statically assign headers
    lldp_nei_headers = [
        "Local Hostname",
        "Local Port",
        "Remote Hostname",
        "Remote Port",
    ]
    # Write headers on the top line of the file
    lldp_nei_ws.append(lldp_nei_headers)
    # Initialize Nornir and define the inventory variables.
    nr = InitNornir(
        inventory={
            "options": {
                "host_file": "inventory/hosts.yaml",
                "group_file": "inventory/groups.yaml",
                "defaults_file": "inventory/defaults.yaml",
            }
        }
    )
    """
    The following block of code assigns a filter based on platform to a variable. This variable is used later on to
    apply logic in for loops
    """
    ios_devices = nr.filter(platform="ios")
    junos_devices = nr.filter(platform="junos")
    eos_devices = nr.filter(platform="eos")
    nxos_devices = nr.filter(platform="nxos")
    iosxr_devices = nr.filter(platform="iosxr")
    """
    Executing the get_interfaces task for each platform so the results
    can be parsed and saved to a spreadsheet
    """
    ios_interfaces = ios_devices.run(name="Processing interfaces", task=get_interfaces)
    junos_interfaces = junos_devices.run(
        name="Processing interfaces", task=get_interfaces
    )
    eos_interfaces = eos_devices.run(name="Processing interfaces", task=get_interfaces)
    nxos_interfaces = nxos_devices.run(
        name="Processing interfaces", task=get_interfaces
    )
    iosxr_interfaces = iosxr_devices.run(
        name="Processing interfaces", task=get_interfaces
    )

    # IOS Platform Block
    for host, task_results in ios_interfaces.items():
        ints = "interfaces"
        get_interfaces_result = task_results[1].result
        interface_name_result = get_interfaces_result[ints]
        # Empty list which will be appended to in for loop
        int_list = []
        # For loop to retrieve the list of interfaces
        for entry in interface_name_result:
            # Append entries to the int_list list
            int_list.append(entry)
        # For loop to loop through list of interfaces and extract interface values
        for int in int_list:
            # Assign individual interface entry to a variable
            final_int = interface_name_result[int]
            # Extract the interface description and assign to a variable
            int_desc_result = final_int["description"]
            # Extract the interface state and assign to a variable
            int_up_result = final_int["is_up"]
            # Extract the whether the interface is enabled and assign to a variable
            int_enable_result = final_int["is_enabled"]
            print(int)
            print(int_desc_result)
            print(int_up_result)
            print(int_enable_result)
            line = [host, int, int_desc_result, int_up_result, int_enable_result]
            # Write values to file
            interfaces_ws.append(line)
    # JUNOS Platform Block
    for host, task_results in junos_interfaces.items():
        ints = "interfaces"
        get_interfaces_result = task_results[1].result
        interface_name_result = get_interfaces_result[ints]
        # Empty list which will be appended to in for loop
        int_list = []
        # For loop to retrieve the list of interfaces
        for entry in interface_name_result:
            # Append entries to the int_list list
            int_list.append(entry)
        # For loop to loop through list of interfaces and extract interface values
        for int in int_list:
            # Assign individual interface entry to a variable
            final_int = interface_name_result[int]
            # Extract the interface description and assign to a variable
            int_desc_result = final_int["description"]
            # Extract the interface state and assign to a variable
            int_up_result = final_int["is_up"]
            # Extract the whether the interface is enabled and assign to a variable
            int_enable_result = final_int["is_enabled"]
            print(int)
            print(int_desc_result)
            print(int_up_result)
            print(int_enable_result)
            line = [host, int, int_desc_result, int_up_result, int_enable_result]
            # Write values to file
            interfaces_ws.append(line)
    # EOS Platform Block
    for host, task_results in eos_interfaces.items():
        ints = "interfaces"
        get_interfaces_result = task_results[1].result
        interface_name_result = get_interfaces_result[ints]
        # Empty list which will be appended to in for loop
        int_list = []
        # For loop to retrieve the list of interfaces
        for entry in interface_name_result:
            # Append entries to the int_list list
            int_list.append(entry)
        # For loop to loop through list of interfaces and extract interface values
        for int in int_list:
            # Assign individual interface entry to a variable
            final_int = interface_name_result[int]
            # Extract the interface description and assign to a variable
            int_desc_result = final_int["description"]
            # Extract the interface state and assign to a variable
            int_up_result = final_int["is_up"]
            # Extract the whether the interface is enabled and assign to a variable
            int_enable_result = final_int["is_enabled"]
            print(int)
            print(int_desc_result)
            print(int_up_result)
            print(int_enable_result)
            line = [host, int, int_desc_result, int_up_result, int_enable_result]
            # Write values to file
            interfaces_ws.append(line)
    # NXOS Platform Block
    for host, task_results in nxos_interfaces.items():
        ints = "interfaces"
        get_interfaces_result = task_results[1].result
        interface_name_result = get_interfaces_result[ints]
        # Empty list which will be appended to in for loop
        int_list = []
        # For loop to retrieve the list of interfaces
        for entry in interface_name_result:
            # Append entries to the int_list list
            int_list.append(entry)
        # For loop to loop through list of interfaces and extract interface values
        for int in int_list:
            # Assign individual interface entry to a variable
            final_int = interface_name_result[int]
            # Extract the interface description and assign to a variable
            int_desc_result = final_int["description"]
            # Extract the interface state and assign to a variable
            int_up_result = final_int["is_up"]
            # Extract the whether the interface is enabled and assign to a variable
            int_enable_result = final_int["is_enabled"]
            print(int)
            print(int_desc_result)
            print(int_up_result)
            print(int_enable_result)
            line = [host, int, int_desc_result, int_up_result, int_enable_result]
            # Write values to file
            interfaces_ws.append(line)
    # IOSXR Platform Block
    for host, task_results in iosxr_interfaces.items():
        ints = "interfaces"
        get_interfaces_result = task_results[1].result
        interface_name_result = get_interfaces_result[ints]
        # Empty list which will be appended to in for loop
        int_list = []
        # For loop to retrieve the list of interfaces
        for entry in interface_name_result:
            # Append entries to the int_list list
            int_list.append(entry)
        # For loop to loop through list of interfaces and extract interface values
        for int in int_list:
            # Assign individual interface entry to a variable
            final_int = interface_name_result[int]
            # Extract the interface description and assign to a variable
            int_desc_result = final_int["description"]
            # Extract the interface state and assign to a variable
            int_up_result = final_int["is_up"]
            # Extract the whether the interface is enabled and assign to a variable
            int_enable_result = final_int["is_enabled"]
            print(int)
            print(int_desc_result)
            print(int_up_result)
            print(int_enable_result)
            line = [host, int, int_desc_result, int_up_result, int_enable_result]
            # Write values to file
            interfaces_ws.append(line)

    ios_facts = ios_devices.run(name="Running a bunch of tasks", task=get_facts)
    junos_facts = junos_devices.run(name="Running a bunch of tasks", task=get_facts)
    eos_facts = eos_devices.run(name="Running a bunch of tasks", task=get_facts)
    nxos_facts = nxos_devices.run(name="Running a bunch of tasks", task=get_facts)
    iosxr_facts = iosxr_devices.run(name="Running a bunch of tasks", task=get_facts)
    # IOS Platform Block
    for host, task_results in ios_facts.items():
        grouped_result = task_results.result
        get_facts_result = task_results[1].result
        vendor_result = get_facts_result["facts"]["vendor"]
        model_result = get_facts_result["facts"]["model"]
        version_result = get_facts_result["facts"]["os_version"]
        ser_num_result = get_facts_result["facts"]["serial_number"]
        uptime_result = get_facts_result["facts"]["uptime"]
        line = [
            host,
            vendor_result,
            model_result,
            version_result,
            ser_num_result,
            uptime_result,
        ]
        # Debug print
        # print(line)
        # Write values to file
        facts_ws.append(line)
        # print(f"{host}: {get_facts_result['facts']['os_version']}")
        print(str(version_result))
        print(str(vendor_result))
        print(str(ser_num_result))
        print(str(model_result))
        print(str(get_facts_result))
        print(str(uptime_result))
        # print(f"---> {grouped_result}")
    # JUNOS Platform Block
    for host, task_results in junos_facts.items():
        grouped_result = task_results.result
        get_facts_result = task_results[1].result
        vendor_result = get_facts_result["facts"]["vendor"]
        model_result = get_facts_result["facts"]["model"]
        version_result = get_facts_result["facts"]["os_version"]
        ser_num_result = get_facts_result["facts"]["serial_number"]
        uptime_result = get_facts_result["facts"]["uptime"]
        line = [
            host,
            vendor_result,
            model_result,
            version_result,
            ser_num_result,
            uptime_result,
        ]
        # Debug print
        # print(line)
        # Write values to facts worksheet
        facts_ws.append(line)
        # print(f"{host}: {get_facts_result['facts']['os_version']}")
        print(str(version_result))
        print(str(vendor_result))
        print(str(ser_num_result))
        print(str(model_result))
        print(str(get_facts_result))
        print(str(uptime_result))
        # print(f"---> {grouped_result}")
    # EOS Platform Block
    for host, task_results in eos_facts.items():
        grouped_result = task_results.result
        get_facts_result = task_results[1].result
        vendor_result = get_facts_result["facts"]["vendor"]
        model_result = get_facts_result["facts"]["model"]
        version_result = get_facts_result["facts"]["os_version"]
        ser_num_result = get_facts_result["facts"]["serial_number"]
        uptime_result = get_facts_result["facts"]["uptime"]
        line = [
            host,
            vendor_result,
            model_result,
            version_result,
            ser_num_result,
            uptime_result,
        ]
        # Debug print
        # print(line)
        # Write values to facts worksheet
        facts_ws.append(line)
        # print(f"{host}: {get_facts_result['facts']['os_version']}")
        print(str(version_result))
        print(str(vendor_result))
        print(str(ser_num_result))
        print(str(model_result))
        print(str(get_facts_result))
        print(str(uptime_result))
        # print(f"---> {grouped_result}")
    # NXOS Platform Block
    for host, task_results in nxos_facts.items():
        grouped_result = task_results.result
        get_facts_result = task_results[1].result
        vendor_result = get_facts_result["facts"]["vendor"]
        model_result = get_facts_result["facts"]["model"]
        version_result = get_facts_result["facts"]["os_version"]
        ser_num_result = get_facts_result["facts"]["serial_number"]
        uptime_result = get_facts_result["facts"]["uptime"]
        line = [
            host,
            vendor_result,
            model_result,
            version_result,
            ser_num_result,
            uptime_result,
        ]
        # Debug print
        # print(line)
        # Write values to facts worksheet
        facts_ws.append(line)
        # print(f"{host}: {get_facts_result['facts']['os_version']}")
        print(str(version_result))
        print(str(vendor_result))
        print(str(ser_num_result))
        print(str(model_result))
        print(str(get_facts_result))
        print(str(uptime_result))
        # print(f"---> {grouped_result}")
    # IOSXR Platform Block
    for host, task_results in iosxr_facts.items():
        grouped_result = task_results.result
        get_facts_result = task_results[1].result
        vendor_result = get_facts_result["facts"]["vendor"]
        model_result = get_facts_result["facts"]["model"]
        version_result = get_facts_result["facts"]["os_version"]
        ser_num_result = get_facts_result["facts"]["serial_number"]
        uptime_result = get_facts_result["facts"]["uptime"]
        line = [
            host,
            vendor_result,
            model_result,
            version_result,
            ser_num_result,
            uptime_result,
        ]
        # Debug print
        # print(line)
        # Write values to facts worksheet
        facts_ws.append(line)
        # print(f"{host}: {get_facts_result['facts']['os_version']}")
        print(str(version_result))
        print(str(vendor_result))
        print(str(ser_num_result))
        print(str(model_result))
        # print(str(get_facts_result))
        print(str(uptime_result))
        # print(f"---> {grouped_result}")
    """
    Executing the get_interfaces_ip task for each platform so the results
    can be parsed and saved to a spreadsheet
    """
    ios_interfaces_ip = ios_devices.run(
        name="Processing interface IP addresses", task=get_interfaces_ip
    )
    junos_interfaces_ip = junos_devices.run(
        name="Processing interface IP addresses", task=get_interfaces_ip
    )
    eos_interfaces_ip = eos_devices.run(
        name="Processing interface IP addresses", task=get_interfaces_ip
    )
    nxos_interfaces_ip = nxos_devices.run(
        name="Processing interface IP addresses", task=get_interfaces_ip
    )
    iosxr_interfaces_ip = iosxr_devices.run(
        name="Processing interface IP addresses", task=get_interfaces_ip
    )
    # IOS Platform Block
    for host, task_results in ios_interfaces_ip.items():
        # Gather results from task
        get_interfaces_ip_result = task_results[1].result
        # Filter the results
        interface_ip_name_result = get_interfaces_ip_result["interfaces_ip"]
        # Empty list which will be appended to in for loop
        int_ip_list = []
        # For loop to retrieve the list of interfaces
        for entry in interface_ip_name_result:
            # Append entries to the int_list list
            int_ip_list.append(entry)
            # Debug print
            # print(int_ip_list)
        # For loop to loop through list of IPv4 interfaces and extract interface_ip values
        for int_ip in int_ip_list:
            # Assign individual interface entry to a variable
            final_int_ip = interface_ip_name_result[int_ip]
            # Assign IPv4 address to a variable
            final_int_ip_addr = final_int_ip["ipv4"]
            # Debug print
            # print(int_ip)
            # For loop to extract single IPv4 address
            for ip in final_int_ip_addr.items():
                # Assign IPv4 address to a variable
                ip_address = ip[0]
                # Debug print
                # print(ip_address)
                # For loop to extract prefix length from prefix_length variable
                for key, prefix_length in ip[1].items():
                    # Print must be left on or for loop isn't activated.
                    print("Extracting prefix length")
                    # Debug print
                    # print(prefix_length)
            # Debug prints
            print(host)
            print(int_ip)
            print(ip_address)
            print(prefix_length)
            # Append results to a line to be saved to the workbook
            line = [host, int_ip, str(ip_address), str(prefix_length)]
            # Save values to row in workbook
            interfaces_ip_ws.append(line)
    # JUNOS Platform Block
    for host, task_results in junos_interfaces_ip.items():
        # Gather results from task
        get_interfaces_ip_result = task_results[1].result
        # Filter the results
        interface_ip_name_result = get_interfaces_ip_result["interfaces_ip"]
        # Empty list which will be appended to in for loop
        int_ip_list = []
        # For loop to retrieve the list of interfaces
        for entry in interface_ip_name_result:
            # Append entries to the int_list list
            int_ip_list.append(entry)
            # Debug print
            # print(int_ip_list)
        # For loop to loop through list of IPv4 interfaces and extract interface_ip values
        for int_ip in int_ip_list:
            # Assign individual interface entry to a variable
            final_int_ip = interface_ip_name_result[int_ip]
            # Assign IPv4 address to a variable
            final_int_ip_addr = final_int_ip["ipv4"]
            # Debug print
            # print(int_ip)
            # For loop to extract single IPv4 address
            for ip in final_int_ip_addr.items():
                # Assign IPv4 address to a variable
                ip_address = ip[0]
                # Debug print
                # print(ip_address)
                # For loop to extract prefix length from prefix_length variable
                for key, prefix_length in ip[1].items():
                    # Print must be left on or for loop isn't activated.
                    print("Extracting prefix length")
                    # Debug print
                    # print(prefix_length)
            # Debug prints
            print(host)
            print(int_ip)
            print(ip_address)
            print(prefix_length)
            # Append results to a line to be saved to the workbook
            line = [host, int_ip, str(ip_address), str(prefix_length)]
            # Save values to row in workbook
            interfaces_ip_ws.append(line)
    # EOS Platform Block
    for host, task_results in eos_interfaces_ip.items():
        # Gather results from task
        get_interfaces_ip_result = task_results[1].result
        # Filter the results
        interface_ip_name_result = get_interfaces_ip_result["interfaces_ip"]
        # Empty list which will be appended to in for loop
        int_ip_list = []
        # For loop to retrieve the list of interfaces
        for entry in interface_ip_name_result:
            # Append entries to the int_list list
            int_ip_list.append(entry)
            # Debug print
            # print(int_ip_list)
        # For loop to loop through list of IPv4 interfaces and extract interface_ip values
        for int_ip in int_ip_list:
            # Assign individual interface entry to a variable
            final_int_ip = interface_ip_name_result[int_ip]
            # Assign IPv4 address to a variable
            final_int_ip_addr = final_int_ip["ipv4"]
            # Debug print
            # print(int_ip)
            # For loop to extract single IPv4 address
            for ip in final_int_ip_addr.items():
                # Assign IPv4 address to a variable
                ip_address = ip[0]
                # Debug print
                # print(ip_address)
                # For loop to extract prefix length from prefix_length variable
                for key, prefix_length in ip[1].items():
                    # Print must be left on or for loop isn't activated.
                    print("Extracting prefix length")
                    # Debug print
                    # print(prefix_length)
            # Debug prints
            print(host)
            print(int_ip)
            print(ip_address)
            print(prefix_length)
            # Append results to a line to be saved to the workbook
            line = [host, int_ip, str(ip_address), str(prefix_length)]
            # Save values to row in workbook
            interfaces_ip_ws.append(line)
    # NXOS Platform Block
    for host, task_results in nxos_interfaces_ip.items():
        # Gather results from task
        get_interfaces_ip_result = task_results[1].result
        # Filter the results
        interface_ip_name_result = get_interfaces_ip_result["interfaces_ip"]
        # Empty list which will be appended to in for loop
        int_ip_list = []
        # For loop to retrieve the list of interfaces
        for entry in interface_ip_name_result:
            # Append entries to the int_list list
            int_ip_list.append(entry)
            # Debug print
            # print(int_ip_list)
        # For loop to loop through list of IPv4 interfaces and extract interface_ip values
        for int_ip in int_ip_list:
            # Assign individual interface entry to a variable
            final_int_ip = interface_ip_name_result[int_ip]
            # Assign IPv4 address to a variable
            final_int_ip_addr = final_int_ip["ipv4"]
            # Debug print
            # print(int_ip)
            # For loop to extract single IPv4 address
            for ip in final_int_ip_addr.items():
                # Assign IPv4 address to a variable
                ip_address = ip[0]
                # Debug print
                # print(ip_address)
                # For loop to extract prefix length from prefix_length variable
                for key, prefix_length in ip[1].items():
                    # Print must be left on or for loop isn't activated.
                    print("Extracting prefix length")
                    # Debug print
                    # print(prefix_length)
            # Debug prints
            print(host)
            print(int_ip)
            print(ip_address)
            print(prefix_length)
            # Append results to a line to be saved to the workbook
            line = [host, int_ip, str(ip_address), str(prefix_length)]
            # Save values to row in workbook
            interfaces_ip_ws.append(line)
    # IOSXR Platform Block
    for host, task_results in iosxr_interfaces_ip.items():
        # Gather results from task
        get_interfaces_ip_result = task_results[1].result
        # Filter the results
        interface_ip_name_result = get_interfaces_ip_result["interfaces_ip"]
        # Empty list which will be appended to in for loop
        int_ip_list = []
        # For loop to retrieve the list of interfaces
        for entry in interface_ip_name_result:
            # Append entries to the int_list list
            int_ip_list.append(entry)
            # Debug print
            # print(int_ip_list)
        # For loop to loop through list of IPv4 interfaces and extract interface_ip values
        for int_ip in int_ip_list:
            # Assign individual interface entry to a variable
            final_int_ip = interface_ip_name_result[int_ip]
            # Assign IPv4 address to a variable
            final_int_ip_addr = final_int_ip["ipv4"]
            # Debug print
            # print(int_ip)
            # For loop to extract single IPv4 address
            for ip in final_int_ip_addr.items():
                # Assign IPv4 address to a variable
                ip_address = ip[0]
                # Debug print
                # print(ip_address)
                # For loop to extract prefix length from prefix_length variable
                for key, prefix_length in ip[1].items():
                    # Print must be left on or for loop isn't activated.
                    print("Extracting prefix length")
                    # Debug print
                    # print(prefix_length)
            # Debug prints
            print(host)
            print(int_ip)
            print(ip_address)
            print(prefix_length)
            # Append results to a line to be saved to the workbook
            line = [host, int_ip, str(ip_address), str(prefix_length)]
            # Save values to row in workbook
            interfaces_ip_ws.append(line)
    """
    Executing the get_lldp_neighbors task for each platform so the results
    can be parsed and saved to a spreadsheet
    """
    ios_lldp = ios_devices.run(
        name="Processing LLDP neighbors", task=get_lldp_neighbors
    )
    junos_lldp = junos_devices.run(
        name="Processing LLDP neighbors", task=get_lldp_neighbors
    )
    eos_lldp = eos_devices.run(
        name="Processing LLDP neighbors", task=get_lldp_neighbors
    )
    nxos_lldp = nxos_devices.run(
        name="Processing LLDP neighbors", task=get_lldp_neighbors
    )
    iosxr_lldp = iosxr_devices.run(
        name="Processing LLDP neighbors", task=get_lldp_neighbors
    )
    # IOS Platform Block
    for host, task_results in ios_lldp.items():
        lldp_nei_result = task_results[1].result
        print(lldp_nei_result)
        # Filter the results
        lldp_nei_name_result = lldp_nei_result["lldp_neighbors"]
        print(lldp_nei_name_result)
        # Empty list which will be appended to in for loop
        neighbor_list = []
        # For loop to retrieve the list of interfaces
        for entry in lldp_nei_name_result:
            # Append entries to the neighbor_list list
            neighbor_list.append(entry)
            print(neighbor_list)
        for local_int in neighbor_list:
            print(local_int)
            remote_port = lldp_nei_name_result[local_int][0]["port"]
            print("remote port")
            print(remote_port)
            remote_hostname = lldp_nei_name_result[local_int][0]["hostname"]
            print("remote hostname")
            print(remote_hostname)
            line = [host, local_int, remote_hostname, remote_port]
            # Write values to file
            lldp_nei_ws.append(line)
    # JUNOS Platform Block
    for host, task_results in junos_lldp.items():
        lldp_nei_result = task_results[1].result
        print(lldp_nei_result)
        # Filter the results
        lldp_nei_name_result = lldp_nei_result["lldp_neighbors"]
        print(lldp_nei_name_result)
        # Empty list which will be appended to in for loop
        neighbor_list = []
        # For loop to retrieve the list of interfaces
        for entry in lldp_nei_name_result:
            # Append entries to the neighbor_list list
            neighbor_list.append(entry)
            print(neighbor_list)
        for local_int in neighbor_list:
            print(local_int)
            remote_port = lldp_nei_name_result[local_int][0]["port"]
            print("remote port")
            print(remote_port)
            remote_hostname = lldp_nei_name_result[local_int][0]["hostname"]
            print("remote hostname")
            print(remote_hostname)
            line = [host, local_int, remote_hostname, remote_port]
            # Write values to file
            lldp_nei_ws.append(line)
    # EOS Platform Block
    for host, task_results in eos_lldp.items():
        lldp_nei_result = task_results[1].result
        print(lldp_nei_result)
        # Filter the results
        lldp_nei_name_result = lldp_nei_result["lldp_neighbors"]
        print(lldp_nei_name_result)
        # Empty list which will be appended to in for loop
        neighbor_list = []
        # For loop to retrieve the list of interfaces
        for entry in lldp_nei_name_result:
            # Append entries to the neighbor_list list
            neighbor_list.append(entry)
            print(neighbor_list)
        for local_int in neighbor_list:
            print(local_int)
            remote_port = lldp_nei_name_result[local_int][0]["port"]
            print("remote port")
            print(remote_port)
            remote_hostname = lldp_nei_name_result[local_int][0]["hostname"]
            print("remote hostname")
            print(remote_hostname)
            line = [host, local_int, remote_hostname, remote_port]
            # Write values to file
            lldp_nei_ws.append(line)
    # NXOS Platform Block
    for host, task_results in nxos_lldp.items():
        lldp_nei_result = task_results[1].result
        print(lldp_nei_result)
        # Filter the results
        lldp_nei_name_result = lldp_nei_result["lldp_neighbors"]
        print(lldp_nei_name_result)
        # Empty list which will be appended to in for loop
        neighbor_list = []
        # For loop to retrieve the list of interfaces
        for entry in lldp_nei_name_result:
            # Append entries to the neighbor_list list
            neighbor_list.append(entry)
            print(neighbor_list)
        for local_int in neighbor_list:
            print(local_int)
            remote_port = lldp_nei_name_result[local_int][0]["port"]
            print("remote port")
            print(remote_port)
            remote_hostname = lldp_nei_name_result[local_int][0]["hostname"]
            print("remote hostname")
            print(remote_hostname)
            line = [host, local_int, remote_hostname, remote_port]
            # Write values to file
            lldp_nei_ws.append(line)
    # IOSXR Platform Block
    for host, task_results in iosxr_lldp.items():
        lldp_nei_result = task_results[1].result
        print(lldp_nei_result)
        # Filter the results
        lldp_nei_name_result = lldp_nei_result["lldp_neighbors"]
        print(lldp_nei_name_result)
        # Empty list which will be appended to in for loop
        neighbor_list = []
        # For loop to retrieve the list of interfaces
        for entry in lldp_nei_name_result:
            # Append entries to the neighbor_list list
            neighbor_list.append(entry)
            print(neighbor_list)
        for local_int in neighbor_list:
            print(local_int)
            remote_port = lldp_nei_name_result[local_int][0]["port"]
            print("remote port")
            print(remote_port)
            remote_hostname = lldp_nei_name_result[local_int][0]["hostname"]
            print("remote hostname")
            print(remote_hostname)
            line = [host, local_int, remote_hostname, remote_port]
            # Write values to file
            lldp_nei_ws.append(line)


def process_functions(wb):
    main_collector(wb)


def create_workbook():
    # Setup workbook parameters
    wb = openpyxl.Workbook()
    # Execute program
    process_functions(wb)
    customer_name = "Customer"
    time_now = (dt.datetime.now()).strftime("%d-%m-%H-%M")
    wb_name = "Diagnostics-" + customer_name + "-2019-" + time_now + ".xlsx"
    print(wb_name)
    wb.save(wb_name)


create_workbook()
